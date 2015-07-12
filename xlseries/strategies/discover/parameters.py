#!/usr/bin/env python
# -*- coding: utf-8 -*-
import json
import pprint
from openpyxl import Workbook
from openpyxl.cell import column_index_from_string

from xlseries.utils.xl_methods import xl_coordinates_range, consecutive_cells

"""
parameters

This module contains the parameters object used by ParameterDiscovery strategy
and all the secondary ones used for it.
"""


# EXCEPTIONS
class InvalidParameter(ValueError):

    """Raised when a parameter is of a non valid value."""

    def __init__(self, param_name, value, valid_values):
        msg = u"""{value} {value_type} is not a valid value for {param_name}
                   parameter. The valid values are {valid_values}.
        """.format(param_name=param_name, value=value,
                   valid_values=valid_values, value_type=type(value))
        super(InvalidParameter, self).__init__(msg)


class CriticalParameterMissing(Exception):

    """Raised when a critical parameter is not provided by the user."""

    def __init__(self, param_name):
        msg = u"{param_name} is a critical parameter. It has to be " + \
            "provided by the user."
        super(CriticalParameterMissing, self).__init__(
            msg.format(param_name=param_name))


class InputParametersNotRecognized(Exception):

    """Raised when parameters obj passed by the user are not recognized."""

    def __init__(self, params):
        msg = u"User input\n{params}\n{params_type} not recognized." + \
            "".format(params=params, params_type=type(params))
        super(InputParametersNotRecognized, self).__init__(msg)


# CLASS
class Parameters(object):

    """Object that collects input parameters from parsing strategies."""

    # this is a complete list of the parameters (of all kinds)
    VALID_VALUES = {
        # general
        "alignment": [u"vertical", u"horizontal"],

        # headers
        "series_names": [str, unicode, None],
        "headers_coord": [str, unicode],

        # data
        "data_starts": [int],
        "data_ends": [int, None],
        "continuity": [True, False],
        "blank_rows": [True, False],
        "missings": [True, False],
        "missing_value": [],

        # time
        "time_alignment": [-1, 0, 1],
        "time_multicolumn": [True, False],
        "time_header_coord": [str, unicode, list],
        "time_composed": [True, False],
        "frequency": ["Y", "Q", "M", "W", "D"]
    }

    # critical values and some template values, as example
    CRITICAL = {
        "time_header_coord": "A1",
        "headers_coord": ["B1", "C1", "E1-G1"],
        "data_starts": 2,
        "frequency": "M"
    }

    # critical values must be provided, they don't have defaults
    # all non critical MUST have default values in this variable
    DEFAULT_VALUES = {
        "time_alignment": 0,
        "alignment": u"vertical",
        "continuity": True,
        "blank_rows": False,
        "missings": False,
        "time_composed": False,
        "time_multicolumn": False,
        "missing_value": [None, "-", "...", ".", ""],
        "data_ends": None,
        "series_names": None
    }

    # order in which default values are most common in xl time series
    LIKELINESS_ORDER = ["time_alignment", "alignment", "continuity",
                        "blank_rows", "missings", "time_composed",
                        "time_multicolumn"]

    # parameters that will be guessed in this class, if missing
    GUESSED = ["time_multicolumn"]

    # parameters that don't need to be specified
    OPTIONAL = ["series_names", "data_ends"]

    # parameters whose default value will be used, if missing
    USE_DEFAULT = ["time_alignment", "missing_value"]

    # auxiliar way to reckon a Parameters object passed to constructor
    TYPE_PARAMETERS = "<class 'xlseries.strategies.discover.parameters.Parameters'>"

    def __init__(self, params_input=None):

        # general
        self.alignment = None

        # name
        self.series_names = None
        self.headers_coord = None

        # data
        self.data_starts = None
        self.data_ends = None
        self.continuity = None
        self.blank_rows = None
        self.missings = None
        self.missing_value = None

        # time
        self.time_alignment = None
        self.time_multicolumn = None
        self.time_header_coord = None
        self.time_composed = None
        self.frequency = None

        if params_input:
            built_params = self._build(self._get_params_dict(params_input))

            for param_name in self.VALID_VALUES:
                if param_name in built_params:
                    setattr(self, param_name, built_params[param_name])
                else:
                    setattr(self, param_name, None)

    @classmethod
    def _get_params_dict(cls, params_input):
        """Return the user input parameters as a dictionary, if possible."""

        if type(params_input) == dict:
            return params_input

        elif (type(params_input) == Parameters or
              unicode(type(params_input)) == cls.TYPE_PARAMETERS):
            return params_input.__dict__

        elif ((type(params_input) == str or type(params_input) == unicode) and
              params_input[-5:] == ".json"):
            with open(params_input) as f:
                return json.load(f)

        else:
            raise InputParametersNotRecognized(params_input)

    @classmethod
    def _build(cls, params_dict):
        """Sanitize and build a complete parameter dict.

        Args:
            params_dict (dict): A parameters dictionary passed by the user
                that has to be curated, validated and completed.
        Returns:
            dict: A complete parameters dict ready to be loaded into
                Parameters instance attributes.
        """

        cls._check_has_critical(params_dict, cls.CRITICAL, cls.VALID_VALUES)

        # curate frequency capitalization
        if type(params_dict["frequency"]) == list:
            params_dict["frequency"] = [i.upper() for i in
                                        params_dict["frequency"]]
        else:
            params_dict["frequency"] = params_dict["frequency"].upper()

        cls._validate_parameters(params_dict, cls.VALID_VALUES)

        params_def = cls._missings_to_default(params_dict, cls.USE_DEFAULT,
                                              cls.VALID_VALUES,
                                              cls.DEFAULT_VALUES)

        # convert ranges of headers (eg. "B8-B28") in lists
        params_def["headers_coord"] = cls._unpack_header_ranges(
            params_def["headers_coord"])
        params_def["time_header_coord"] = cls._unpack_header_ranges(
            params_def["time_header_coord"])

        # import pdb
        # pdb.set_trace()
        cls._check_consistency(params_def)

        # guess parameters based on other parameters
        num_series = cls._get_num_series(params_def)
        if ("time_multicolumn" not in params_def or
                not params_def["time_multicolumn"]):
            params_def["time_multicolumn"] = cls._guess_time_multicolumn(
                params_def["time_header_coord"], num_series)

        if "alignment" not in params_def or not params_def["alignment"]:
            params_def["alignment"] = cls._guess_alignment(
                params_def["headers_coord"])

        # apply single provided parameters to all series
        for param_name in params_def.keys():
            params_def[param_name] = cls._apply_to_all(
                param_name, params_def[param_name], num_series, params_def,
                cls.VALID_VALUES[param_name])

        # apply Nones to optional parameters
        for param_name in cls.OPTIONAL:
            if param_name not in params_def or not params_def[param_name]:
                params_def[param_name] = cls._apply_to_all(
                    param_name, cls.DEFAULT_VALUES[param_name], num_series,
                    params_def, cls.VALID_VALUES[param_name])

        return params_def

    def __repr__(self):
        return pprint.pformat(self.__dict__)

    def __getitem__(self, item):

        if type(item) == int:
            return self.get_series_params(item)

        else:
            return self.__getattribute__(item)

    def __setitem__(self, param_name, param_value):
        num_series = self._get_num_series(self.__dict__)
        if self._valid_param_list(param_name, param_value, num_series):
            self.__dict__[param_name] = param_value

        else:
            if not self._valid_param_value(param_value,
                                           self.VALID_VALUES[param_name]):
                raise InvalidParameter(param_name, param_value,
                                       self.VALID_VALUES[param_name])

            self.__dict__[param_name] = self._apply_to_all(
                param_name, param_value, self._get_num_series(
                    self.__dict__), self,
                self.VALID_VALUES[param_name])

    def __iter__(self):
        for param in self.__dict__:
            yield param

    def __eq__(self, other):
        for key in self:
            if self[key] != other[key]:
                return False

        for key in other:
            if self[key] != other[key]:
                return False

        return True

    def __len__(self):
        return self._get_num_series(self.__dict__)

    # PUBLIC
    def get_series_params(self, i_series):
        """Returns parameters for only one series."""

        series_params = Parameters().__dict__

        for param_name in series_params.keys():
            if self[param_name] is not None:
                series_params[param_name] = self[param_name][i_series]

            # missing parameters have to be deleted from slicing so they are
            # not confused with possible valid None values
            else:
                del series_params[param_name]

        return series_params

    def is_complete(self):
        """Check if all the parameters have values (ie. no misssing params)."""
        num_series = self._get_num_series(self.__dict__)

        for param_name in self:
            if param_name == "time_header_coord":
                if (param_name not in self.OPTIONAL and
                    (self[param_name] is None or
                     len(self[param_name]) < min(2, num_series))):
                    return False

            else:
                if (param_name not in self.OPTIONAL and
                    (self[param_name] is None or
                     len(self[param_name]) != num_series)):
                    return False

        return True

    def get_missings(self):
        """Return the names of parameters that were not passed by the user."""
        missings = []
        for param in self:
            if self._is_missing(param):
                missings.append(param)
        return missings

    def get_non_critical_params(self, differents=True):
        """Return the name of non critical parameters.

        Args:
            differents (bool): If True, will also get non critical parameters
                that have differences across series.
        """
        if not differents:
            return [param for param in self if (self._non_critical(param) and
                                                self._no_differences(param))]
        else:
            return [param for param in self if self._non_critical(param)]

    def num_missings(self):
        """Return the number of missing parameters."""
        return sum((1 for param in self if self._is_missing(param)))

    def remove_non_critical(self, differents=False):
        """Remove all non critical parameters but the ones that use default."""
        [self.remove(i) for i in self.get_non_critical_params(differents) if
         i not in self.USE_DEFAULT and i not in self.GUESSED]

        # restore optionals
        for optional_param in self.OPTIONAL:
            self[optional_param] = None

    def remove(self, param):
        """Remove a parameters setting it to 'missing'."""
        self.__dict__[param] = None

    @classmethod
    def get_critical_params_template(cls):
        """Return a template dictionary of critical params."""
        return {param: value for param, value in
                cls.CRITICAL.iteritems() if
                param in cls.CRITICAL}

    @classmethod
    def get_complete_params_template(cls):
        """Return a template dictionary of critical params."""
        return dict(cls.CRITICAL.items() + cls.DEFAULT_VALUES.items())

    def compact_repr(self):
        """Return a dict of the parameters in their compact representation.

        If a parameter is repeated in all the series, show the parameter
        without repeating."""

        params_compact = {}
        for param_name in self:
            if (not self._is_missing(param_name) and
                    not self._is_optional(param_name) and
                    not self._is_default(param_name)):

                if self._is_repeated(self[param_name]):
                    params_compact[param_name] = self[param_name][0]
                else:
                    params_compact[param_name] = self[param_name]

        return params_compact

    # PRIVATE for build step 1: check critical parameters were passed
    @classmethod
    def _check_has_critical(cls, params_dict, critical, valid_values):
        """Check that a dictionary of parameters has all critical ones."""

        for critical_param in critical:
            if critical_param not in params_dict:
                raise CriticalParameterMissing(critical_param)

            elif (params_dict[critical_param] is None and
                    None not in valid_values[critical_param]):
                raise CriticalParameterMissing(critical_param)

    # PRIVATE for build step 2: validate passed parameter values
    @classmethod
    def _validate_parameters(cls, params_dict, valid_values):
        """Check that all values of the parameters are valid."""

        for param_name, param_value in params_dict.iteritems():

            # if a parameter is not provided, its validity cannot be checked
            if param_value is None:
                continue

            # param value may be passed as unique value or as a complete list
            if type(param_value) == list:
                iter_param_values = param_value
            else:
                iter_param_values = [param_value]

            for value in iter_param_values:
                if param_name == "frequency":
                    if not cls._valid_freq(value, valid_values["frequency"]):
                        raise InvalidParameter(param_name, value)

                else:
                    if not cls._valid_param_value(value,
                                                  valid_values[param_name]):
                        raise InvalidParameter(param_name, value,
                                               valid_values[param_name])

    @classmethod
    def _valid_freq(cls, value, valid_values):
        """Check that a frequency is composed of valid frequency characters."""

        for char in value:
            if char not in valid_values:
                return False
        return True

    @classmethod
    def _valid_param_value(cls, value, valid_values):
        """Check that a value is valid.

        Check against a list of valid values or valid types of values."""

        if not valid_values:
            return True

        for valid_value in valid_values:
            if type(valid_value) == type and type(value) == valid_value:
                return True

            elif value == valid_value:
                return True

        return False

    # PRIVATE for build step 3: set missings to default values, when possible
    @classmethod
    def _missings_to_default(cls, params_dict, use_default, valid_values,
                             default_values):
        """Set missing parameters in the USE_DEFAULT list to their defaults.

        These parameters, when not provided by the user, will use their
        defaults rather than stay missing to be guessed by the package in a
        later stage."""
        params_def = params_dict.copy()

        for param_default in use_default:
            if ((param_default not in params_dict or
                 params_dict[param_default] is None) and
                    None not in valid_values[param_default]):
                params_def[param_default] = default_values[param_default]

        return params_def

    # PRIVATE for build step 4: unpack header cell ranges
    @classmethod
    def _unpack_header_ranges(cls, coord_param):

        if (not coord_param or len(coord_param) == 0 or
                (type(coord_param) != list and coord_param.lower() == "none")):
            return None

        return list(cls._unpack_header_ranges_generator(coord_param))

    @classmethod
    def _unpack_header_ranges_generator(cls, coord_param):

        if type(coord_param) == str or type(coord_param) == unicode:
            if "-" not in coord_param and ":" not in coord_param:
                yield coord_param.upper()
            else:
                if "-" in coord_param:
                    start, end = coord_param.upper().split("-")
                elif ":" in coord_param:
                    start, end = coord_param.upper().split(":")

                for cell in xl_coordinates_range(start, end):
                    yield cell

        elif type(coord_param) == list:
            for elem in coord_param:
                if type(elem) == list:
                    yield list(cls._unpack_header_ranges_generator(elem))
                else:
                    for unpacked in cls._unpack_header_ranges_generator(elem):
                        yield unpacked

    # PRIVATE for build step 5: check consistency
    @classmethod
    def _check_consistency(cls, params_def):

        # check data starts is consistent with headers coordinates
        ws = Workbook().active
        if type(params_def["data_starts"]) == list:
            data_starts = params_def["data_starts"][0]
        else:
            data_starts = params_def["data_starts"]

        if type(params_def["headers_coord"]) == list:
            rows = [ws[coord].row for coord in params_def["headers_coord"]]
            cols = [column_index_from_string(ws[coord].column)
                    for coord in params_def["headers_coord"]]

            probably_vertical, probably_horizontal = None, None
            if "alignment" in params_def:
                alignment = params_def["alignment"]
            else:
                alignment = None

                if len(params_def["headers_coord"]) > 1:
                    probably_vertical = (len(set(rows)) == 1 and
                                         len(set(cols)) == len(cols))
                    probably_horizontal = (len(set(cols)) == 1 and
                                           len(set(rows)) == len(rows))

            if alignment == "vertical" or probably_vertical:
                msg = "Row {} where data starts, must fe after {} where " + \
                    "headers are."
                assert data_starts > rows[0], msg.format(data_starts, rows[0])

            if alignment == "horizontal" or probably_horizontal:
                msg = "Column {} where data starts, must fe after {} where" + \
                    " headers are.".format(data_starts, cols[0])
                assert data_starts > cols[0], msg

    # PRIVATE for build step 6: guess parameters
    @classmethod
    def _guess_time_multicolumn(cls, time_header_coord, num_series):
        """Guess if a time index is multicolumn.

        Based on the number of time_header_coord elements compared with the
        number of series. If they are different and time_header_coord has more
        than one, it will be a multicolumn.
        """

        assert "time_multicolumn" in cls.GUESSED, "time_multicolumn is " + \
            "not in the guessed list."

        if cls._is_repeated(time_header_coord):
            tch = time_header_coord[0]
        else:
            tch = time_header_coord

        if (type(tch) == list and
                len(tch) != num_series):
            return True
        elif (type(tch) == list and
                len(tch) == num_series and
                consecutive_cells(tch)):
            return True
        else:
            return False

    @classmethod
    def _guess_alignment(cls, headers_coord):
        """Guess alignment of series.

        If all the header coords are in the same row is vertical, in the same
        column is horizontal. If less than 4, the headers must be consecutive
        to be able to use this guessing. With >4 non consecutive ones are
        allowed."""
        # import pdb; pdb.set_trace()
        ws = Workbook().active
        if type(headers_coord) != list or len(headers_coord) <= 1:
            return None

        if ((len(headers_coord) < 4 and consecutive_cells(headers_coord)) or
                len(headers_coord) >= 4):
            rows = [ws[coord].row for coord in headers_coord]
            cols = [column_index_from_string(ws[coord].column)
                    for coord in headers_coord]

            if len(set(rows)) == 1 and len(set(cols)) == len(cols):
                return "vertical"

            elif len(set(cols)) == 1 and len(set(rows)) == len(rows):
                return "horizontal"

        return None

    # PRIVATE for build step 7: apply parameters to all series
    @classmethod
    def _apply_to_all(cls, param_name, param_value, num_series, params,
                      valid_values=None):
        """Creates list from single parameter repeating it for every series."""

        if param_name == "time_header_coord":
            return cls._apply_to_all_time_header(param_value, num_series,
                                                 params, valid_values)
        elif param_name == "missing_value":
            return cls._apply_to_all_missing_value(param_value, num_series)

        if (param_value is None and valid_values and None not in valid_values):
            return None

        elif not type(param_value) == list and num_series:
            param_list = [param_value for i in xrange(num_series)]

        elif (type(param_value) == list and len(param_value) == 1 and
                num_series):
            param_list = [param_value[0] for i in xrange(num_series)]

        else:
            param_list = param_value

        return param_list

    @classmethod
    def _apply_to_all_time_header(cls, time_header_coord, num_series, params,
                                  valid_values=None):
        """Creates list from single parameter repeating it for every series."""

        if (type(time_header_coord) == list and
                type(time_header_coord[0]) == list):
            if len(time_header_coord) == num_series:
                return time_header_coord
            else:
                raise ValueError("time_header_coord list of lists has to be" +
                                 " of " + unicode(num_series) + " length.")

        elif type(params["time_multicolumn"]) == list:
            time_multicolumn = params["time_multicolumn"][0]

        else:
            time_multicolumn = params["time_multicolumn"]

        if (not type(time_header_coord) == list or not time_multicolumn):
            return cls._apply_to_all("", time_header_coord, num_series,
                                     params, valid_values)
        else:
            return [time_header_coord for i in xrange(num_series)]

    @classmethod
    def _apply_to_all_missing_value(cls, missing_value, num_series):
        """Creates list from single parameter repeating it for every series."""

        if not type(missing_value) == list:
            return [[missing_value] for i in xrange(num_series)]

        elif (type(missing_value) == list and
              (len(missing_value) == 0 or type(missing_value[0]) != list)):
            return [missing_value for i in xrange(num_series)]

        else:
            msg = "If missing values are specified for every single " + \
                "series, you must specify them for the " + unicode(num_series)
            assert len(missing_value) == num_series, msg
            return missing_value

    # PRIVATE AUXILIAR for public methods or broadly used
    @classmethod
    def _get_num_series(cls, params):
        """Count number of series present in parameters."""

        num_series = None
        for param_name, param_value in params.iteritems():
            if type(param_value) == list and param_name != "missing_value":
                if not num_series or len(param_value) > num_series:
                    num_series = len(param_value)

        return num_series

    def _is_optional(self, param_name):
        """True if parameter is optional and is set to None."""
        if self._is_repeated(self[param_name]):
            return param_name in self.OPTIONAL and self[param_name][0] is None
        else:
            return param_name in self.OPTIONAL and self[param_name] is None

    def _is_default(self, param_name):
        """True if parameter is a USE_DEFAULT and is set to its def value."""
        if self._is_repeated(self[param_name]):
            return (param_name in self.USE_DEFAULT and
                    self.DEFAULT_VALUES[param_name] == self[param_name][0])
        elif not type(param_name) == list:
            return (param_name in self.USE_DEFAULT and
                    self.DEFAULT_VALUES[param_name] == self[param_name])
        else:
            return False

    @classmethod
    def _is_repeated(cls, param_value):
        if not type(param_value) == list:
            return False

        return all(i == param_value[0] for i in param_value)

    def _valid_param_list(self, param_name, param_value, num_series):
        """Return True if param_value is a valid list of parameters for
        param_name."""
        return (type(param_value) == list and
                len(param_value) == num_series and
                all(self._valid_param_value(param,
                                            self.VALID_VALUES[param_name])
                    for param in param_value)
                )

    def _non_critical(self, param):
        """Return True if param is not critical."""
        return param not in self.CRITICAL

    def _no_differences(self, param):
        """Return True if param is the same for all the series."""
        if type(self[param]) == list and type(self[param][0]) != list:
            return len(set(self[param])) == 1
        else:
            return True

    def _is_missing(self, param):
        valid_values = self.VALID_VALUES[param]
        return (self[param] is None and None not in valid_values)
