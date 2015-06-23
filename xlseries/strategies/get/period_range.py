#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
period_range

This module contains strategies to get period ranges from a clean worksheet.

Be aware that every class in this namespace whose name doesn't start with
"Base" will be returned by get_strategies() as a valid strategy to iterate
looking for accepting or refusing certain input.
"""

from __future__ import unicode_literals
from pprint import pprint
from openpyxl.cell import column_index_from_string
import pandas as pd

import xlseries.utils.strategies_helpers


class BaseGetPeriodRangesStrategy(object):

    """Base class for all strategies to get period ranges."""

    # PUBLIC INTERFACE
    @classmethod
    def accepts(cls, ws, freq):
        return cls._accepts(ws, freq)

    @classmethod
    def get_period_ranges(cls, ws, freq, ini_row, time_header_coord, end_row,
                          time_alignement, alignment):
        return cls._get_period_ranges(ws, freq, ini_row, time_header_coord,
                                      end_row, time_alignement, alignment)


class GetPeriodRangesSingleFrequency(BaseGetPeriodRangesStrategy):

    """Get period ranges for time series of a single frequency."""

    @classmethod
    def _accepts(cls, ws, freq):
        return len(freq) == 1

    @classmethod
    def _get_period_ranges(cls, ws, freq, ini, time_header_coord, end,
                           time_alignement, alignment):

        if alignment == "vertical":
            col = column_index_from_string(ws[time_header_coord].column)
            start = ws.cell(row=ini + time_alignement, column=col).value
            end = ws.cell(row=end + time_alignement, column=col).value

        elif alignment == "horizontal":
            row = ws[time_header_coord].row
            start = ws.cell(column=ini + time_alignement, row=row).value
            end = ws.cell(column=end + time_alignement, row=row).value

        else:
            raise Exception("Series alignment must be 'vertical' or " +
                            "'horizontal', not " + repr(alignment))

        return [pd.period_range(start, end, freq=freq)]


class GetPeriodRangesMultifrequency(BaseGetPeriodRangesStrategy):

    """Get period ranges for multifrequency time series."""

    @classmethod
    def _accepts(cls, ws, freq):
        return len(freq) > 1

    @classmethod
    def _get_period_ranges(cls, ws, freq, ini, time_header_coord, end,
                           time_alignement, alignment):

        starts = {f: None for f in freq}
        ends = {f: None for f in freq}

        if type(time_header_coord) == list:
            th_coord = time_header_coord[0]
        else:
            th_coord = time_header_coord

        if alignment == "vertical":
            col = column_index_from_string(ws[th_coord].column)

            # capture starting times
            rows = ws.rows[ini - 1:ini + len(freq) - 1]
            for cell, f in zip((row[col - 1] for row in rows), freq):
                if not starts[f]:
                    starts[f] = cell.value

            # capture ending times
            # calculates if multifreq series stop before a complete cycle
            freq_end = (end - ini + 1) % len(freq)
            if freq_end == 0:
                freq_end = len(freq)

            rows = ws.rows[end - freq_end:end]
            for cell, f in zip((row[col - 1] for row in reversed(rows)),
                               freq[:freq_end][::-1]):
                if not ends[f]:
                    ends[f] = cell.value

        elif alignment == "horizontal":
            row = ws[th_coord].row

            # capture starting times
            cols = ws.columns[ini - 1:ini + len(freq) - 1]
            for cell, f in zip((col[row - 1] for col in cols), freq):
                if not starts[f]:
                    starts[f] = cell.value

            # capture ending times
            # calculates if multifreq series stop before a complete cycle
            freq_end = (end - ini + 1) % len(freq)
            if freq_end == 0:
                freq_end = len(freq)

            cols = ws.columns[end - freq_end:end]
            for cell, f in zip((col[row - 1] for col in reversed(cols)),
                               freq[:freq_end][::-1]):
                if not ends[f]:
                    ends[f] = cell.value

        else:
            raise Exception("Series alignment must be 'vertical' or " +
                            "'horizontal', not " + repr(alignment))

        return [pd.period_range(starts[f], ends[f], freq=f) for f in starts]


def get_strategies():
    return xlseries.utils.strategies_helpers.get_strategies()

if __name__ == '__main__':
    pprint(sorted(xlseries.utils.strategies_helpers.get_strategies_names()))
